"""Verifica que el avance por materia cuenta como acreditadas también las
materias que el alumno cursó del plan companion "no vigentes cursadas por las
cohortes en estudio", no sólo las del plan vigente asignado (riesgo #5 del
diseño). Caso de referencia: alumno 0248323 (titulado), cuya vista de avance
mostraba materias aprobadas como "no acreditadas" antes de la corrección.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import app.db.base as base

DATA = Path(__file__).resolve().parents[2] / "data" / "ejemplos"


@pytest.fixture()
def db():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    base.engine = engine
    base.SessionLocal = SessionLocal

    import app.models  # noqa: F401  (registra los modelos en Base.metadata)
    from app.db import seed as seed_mod

    base.Base.metadata.create_all(bind=engine)
    seed_mod.engine = engine
    seed_mod.SessionLocal = SessionLocal
    seed_mod.seed()

    from app.services import csv_import
    from app.models import Carrera

    session = SessionLocal()
    csv_import.import_plan_estudios(session, (DATA / "plan_estudios.csv").read_bytes(), "plan.csv")
    carrera = session.query(Carrera).first()
    csv_import.import_alumnos(session, (DATA / "alumnos.csv").read_bytes(), "alumnos.csv", carrera.clave)
    csv_import.import_kardex(session, (DATA / "kardex.csv").read_bytes(), "kardex.csv", carrera.clave)
    try:
        yield session
    finally:
        session.close()


def test_materias_del_plan_companion_cuentan_como_acreditadas(db):
    from app.services.trayectoria import obtener_trayectoria, construir_avance

    trayectoria = obtener_trayectoria(db, "0248323")
    data = construir_avance(db, trayectoria)

    aprobadas_en_kardex = {k["materia_cve"] for k in data["kardex"] if k["resultado"] == "aprobado"}
    acreditadas_en_avance = {m["materia_cve"] for m in data["avance_por_materia"] if m["acreditada"]}

    # Toda materia aprobada que exista en algún plan de la carrera debe figurar
    # como acreditada en el avance.
    materias_del_plan = {m["materia_cve"] for m in data["avance_por_materia"]}
    faltantes = (aprobadas_en_kardex & materias_del_plan) - acreditadas_en_avance
    assert not faltantes, f"materias aprobadas no marcadas como acreditadas: {sorted(faltantes)}"

    # Antes de la corrección sólo se contaban ~48 materias (plan vigente); con el
    # plan companion incluido deben superar ese número.
    assert len(acreditadas_en_avance) >= 60


def test_avance_se_mide_contra_creditos_de_pasantia(db):
    from app.services.trayectoria import (
        CREDITOS_REQUERIDOS_PASANTE,
        construir_avance,
        obtener_trayectoria,
    )

    data = construir_avance(db, obtener_trayectoria(db, "0248323"))

    # El avance se mide contra los créditos de pasantía y se topa en 100%.
    assert data["creditos_requeridos"] == CREDITOS_REQUERIDOS_PASANTE
    esperado = min(data["creditos_acreditados"] / CREDITOS_REQUERIDOS_PASANTE, 1.0)
    assert data["porcentaje_avance"] == pytest.approx(esperado)
    assert 0.0 <= data["porcentaje_avance"] <= 1.0




def test_exportar_tabla_genera_xlsx():
    from openpyxl import load_workbook

    from app.services.excel_export import generar_tabla

    buffer = generar_tabla(
        "Aprobación",
        ["Ciclo", "Clave", "% Aprobados"],
        [["2015-2016/I", "0041", 0.95], ["2016-2017/II", "0042", 1.0]],
    )
    wb = load_workbook(buffer)
    ws = wb.active
    assert ws.title == "Aprobación"
    assert [c.value for c in ws[1]] == ["Ciclo", "Clave", "% Aprobados"]
    assert ws.max_row == 3
    assert ws["C2"].value == 0.95
