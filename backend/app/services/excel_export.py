"""Genera un libro de Excel equivalente al archivo original, a partir de
los datos calculados por el motor de indicadores (app.services.indicadores).
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from sqlalchemy.orm import Session, joinedload

from app.models import Carrera, Trayectoria
from app.services import indicadores as svc


def _hoja_resumen(wb: Workbook, db: Session, carrera_clave: str) -> None:
    ws = wb.create_sheet("Resumen")
    encabezados = [
        "Cohorte", "Alumnos Cohorte", "Abandono", "Retención",
        "Titulados", "Pasantes", "Titulados % Cohorte", "Titulados % Retenidos",
        "Egresados", "Egresados % Cohorte", "Egresados % Retenidos",
        "Reprobadas 0", "Reprobadas 1-3", "Reprobadas 4+",
    ]
    ws.append(encabezados)
    for fila in svc.resumen_cohortes(db, carrera_clave):
        ws.append(
            [
                fila["cohorte"], fila["alumnos_cohorte"], fila["abandono"], fila["retencion"],
                fila["titulados"], fila["pasantes"], fila["titulados_pct_cohorte"], fila["titulados_pct_retenidos"],
                fila["egresados"], fila["egresados_pct_cohorte"], fila["egresados_pct_retenidos"],
                fila["reprobadas_0"], fila["reprobadas_1a3"], fila["reprobadas_4mas"],
            ]
        )


def _hoja_aprobacion(wb: Workbook, db: Session, carrera_clave: str) -> None:
    ws = wb.create_sheet("Aprobación")
    ws.append(["Ciclo", "Clave Materia", "Materia", "Inscritos", "% Aprobados Ordinario", "% Aprobados"])
    for fila in svc.aprobacion_por_materia(db, carrera_clave):
        ws.append(
            [
                fila["ciclo"], fila["materia_cve"], fila["materia_nombre"], fila["inscritos"],
                fila["pct_aprobados_ordinario"], fila["pct_aprobados"],
            ]
        )


def _hoja_promedios(wb: Workbook, db: Session, carrera_clave: str) -> None:
    ws = wb.create_sheet("Calificación Promedio Materia")
    ws.append(["Ciclo", "Clave Materia", "Materia", "Promedio", "% Reprobación"])
    for fila in svc.promedios_por_materia(db, carrera_clave):
        ws.append([fila["ciclo"], fila["materia_cve"], fila["materia_nombre"], fila["promedio"], fila["pct_reprobacion"]])


def _hoja_titulacion(wb: Workbook, db: Session, carrera_clave: str) -> None:
    ws = wb.create_sheet("Titulación")
    filas = svc.titulacion_por_cohorte(db, carrera_clave)
    modalidades = sorted({m for f in filas for m in f["modalidades"]})
    ws.append(["Cohorte", *modalidades, "Total"])
    for f in filas:
        ws.append([f["cohorte"], *[f["modalidades"].get(m, 0) for m in modalidades], f["total"]])


def _hoja_egreso(wb: Workbook, db: Session, carrera_clave: str) -> None:
    ws = wb.create_sheet("Egreso")
    ws.append(["Cohorte", "Promedio de Semestres"])
    for f in svc.egreso_por_cohorte(db, carrera_clave):
        ws.append([f["cohorte"], f["promedio_semestres"]])


def _hoja_retencion(wb: Workbook, db: Session, carrera_clave: str) -> None:
    ws = wb.create_sheet("Rezago y Retención")
    matriz = svc.retencion_matriz(db, carrera_clave)
    ciclos = matriz["ciclos"]
    ws.append(["Alumno", "Cohorte", "Situación", "Semestres Cursados", "Rezago Aproximado", *ciclos])
    for fila in matriz["alumnos"]:
        ws.append(
            [
                fila["cve_uaslp"], fila["cohorte"], fila["situacion"],
                fila["semestres_cursados"], fila["rezago_aproximado"],
                *[fila["ciclos"].get(c, 0) for c in ciclos],
            ]
        )


def _hoja_kardex(wb: Workbook, db: Session, carrera: Carrera) -> None:
    ws = wb.create_sheet("Kardex")
    ws.append(["cve_uaslp", "creditos", "cve_materia", "materia", "calificación", "fecha_cal", "tipo_examen", "semestre"])
    trayectorias = (
        db.query(Trayectoria)
        .options(joinedload(Trayectoria.alumno), joinedload(Trayectoria.kardex))
        .filter(Trayectoria.carrera_id == carrera.id)
        .all()
    )
    for t in trayectorias:
        for k in t.kardex:
            calif = k.calificacion_numerica if k.calificacion_numerica is not None else (
                k.codigo_calificacion.clave if k.codigo_calificacion else ""
            )
            ws.append(
                [
                    t.alumno.cve_uaslp, k.creditos, k.materia.cve_materia, k.materia.nombre,
                    calif, k.fecha_cal, k.tipo_examen.clave if k.tipo_examen else "", k.ciclo.nombre,
                ]
            )


def generar_libro(db: Session, carrera_clave: str) -> io.BytesIO:
    carrera = db.query(Carrera).filter(Carrera.clave == carrera_clave.upper()).one_or_none()
    if carrera is None:
        raise ValueError(f"carrera '{carrera_clave}' no encontrada")

    wb = Workbook()
    wb.remove(wb.active)  # quitar la hoja por defecto
    _hoja_kardex(wb, db, carrera)
    _hoja_resumen(wb, db, carrera_clave)
    _hoja_retencion(wb, db, carrera_clave)
    _hoja_aprobacion(wb, db, carrera_clave)
    _hoja_promedios(wb, db, carrera_clave)
    _hoja_titulacion(wb, db, carrera_clave)
    _hoja_egreso(wb, db, carrera_clave)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
